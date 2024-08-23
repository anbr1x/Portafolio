import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import glob
import os
import datetime
import os
import json

def buscar_archivo_mas_antiguo(directorio, cadena):
    archivos_coincidentes = [os.path.join(directorio, nombre_archivo) for nombre_archivo in os.listdir(directorio) if cadena in nombre_archivo]
    if not archivos_coincidentes:
        return None
    else:
        archivo_mas_antiguo = min(archivos_coincidentes, key=os.path.getmtime)
        return archivo_mas_antiguo

def contar_prefijos(lista):
        contador = defaultdict(int)
        for cadena in lista:
            for i in range(1, len(cadena) + 1):
                prefijo = cadena[:i]
                contador[prefijo] += 1
        return contador
def MostCL_prefix(lista : list):
    b = 0
    list_n = []
    list_n1 = {}
    contador = contar_prefijos(lista)
    for key in contador.keys(): 
       ###Aqui iria la nueva condicional### 
        if contador[key] < b:
               list_n.append(key)
        b = contador[key]
    claves = list(contador.keys())
    for clave_dada in list_n:
           indice = claves.index(clave_dada)
           clave_anterior = claves[indice - 1]
           #list_n.append(clave_anterior)
           list_n1[clave_anterior] = contador[clave_anterior]

    return list_n1,contador

def load_TSAP(path:str):
    # Dicc para almacenar los DataFrames
    dfs = {}
    # Itera sobre todos los archivos en el directorio
    for filename in tqdm.tqdm(os.listdir(directorio)):
        file_path = os.path.join(directorio, filename)
        # Verifica si el path es un archivo y no un directorio
        if os.path.isfile(file_path):
            df = pd.read_excel(file_path)  
            # Obtiene el nombre del archivo sin la extensión
            nombre_sin_extension = os.path.splitext(filename)[0]
            # Añade el DataFrame al diccionario
            dfs[nombre_sin_extension] = df
    df_combinado = pd.concat(dfs.values())
    return df_combinado

def load_last_run_date(filename):
    if os.path.exists(filename):
        with open(filename, 'r') as f:
            return datetime.datetime.strptime(json.load(f)['date'], '%Y-%m-%d').date()
    else:
        return None

def save_last_run_date(filename):
    with open(filename, 'w') as f:
        json.dump({'date': datetime.datetime.now().strftime('%Y-%m-%d')}, f)

def process_df(df: pd.DataFrame,clases_unicas: list,index=['NOMPROVEEDOR','ESTADO']):
    """ La funcion recibe un df, y las clases unicas
        Y devuelve el df uniformizado para cada clase
        Llena los valores faltantes con 0"""
    pivot_table = pd.pivot_table(df, values='PEND_FACT_SOLES', index=index, aggfunc=pd.Series.sum) #Agrupo 
    grupos = pivot_table.groupby('NOMPROVEEDOR') #creo un data frame para cada contrata
    # Para cada grupo, crea un nuevo dataframe y guárdalo en un diccionario
    dataframes = {}
    for nombre, datos in grupos: # Creo los df dentro del dic
        dataframes[nombre] = datos
    for key in dataframes.keys(): # para cada df en el dic
        df = dataframes[key].reset_index()
        df1 = df.set_index('ESTADO').reindex(clases_unicas).reset_index() #normalizo, creando filas para todos los ESTADOS
        df1.NOMPROVEEDOR = key  #Relleno los NAN
        df1 = df1.fillna(0)  
        dataframes[key] = df1   #Reescribo los Df para cada clave
    df_concat = pd.concat(dataframes.values(),ignore_index=True) # Compacto todo los df del dic en uno grande 
    pivot_table_1 = pd.pivot_table(df_concat, values='PEND_FACT_SOLES', index=['NOMPROVEEDOR', 'ESTADO'], aggfunc=pd.Series.sum) #agrupo de nuevo
    return pivot_table_1


def calc_diff(df1: pd.DataFrame,df2: pd.DataFrame, index= ['NOMPROVEEDOR', 'ESTADO']) -> pd.DataFrame:
    """ entran 2 dataframes , los proceso , y en base a los df procesados 
        calculo la diferencia en una nueva columna, 
        y añado esta columna de diferencia al df actual,
        DF1 - DF2"""
    clases_unicas = pd.concat([df1['ESTADO'], df2['ESTADO']]).unique() #Creo listas de claves unicas
    df1_proces = process_df(df1,clases_unicas,index)
    df2_proces = process_df(df2,clases_unicas,index)
    
    df_diff = df1_proces - df2_proces
    diff_pivot_table_reset = df_diff.reset_index()
    diff_pivot_table_reset.rename(columns={'PEND_FACT_SOLES': 'DIFERENCIA EN SOLES'}, inplace=True)
    # Merge diff_pivot_table_filled_reset con PrePa_O_EI
    PrePa_O_EI = pd.merge(df1, diff_pivot_table_reset, on=index+['ESTADO'],how='left')    
    return PrePa_O_EI

class Alerts(list):
    def __init__(self,list):
        self.data_dict = {
            list[0]: [],
            list[1]: [],
            list[2]: [],
            list[3]: [],
            list[4]: [],
            list[5]: [],
            list[6]: [],
            list[7]: [],
            list[8]: []
        }


def update_values(df1 , df2, index_name, columns): 
    idx = index_name

    def update(row, column):
        if row[idx] in df2[idx].values:
            idx_value = row[idx]
            if pd.isna(df2.loc[df2[idx] == idx_value, column].values[0]):
                return row[column]
            else:
                return df2.loc[df2[idx] == idx_value, column].values[0]
        else:
            return row[column]

    for column in columns:
        df1.loc[:, column] = df1.apply(lambda row: update(row, column), axis=1)
    return df1

def update_values_optimized(df1, df2, index_name, columns):
    """ ACtualiza columna de un dataframe con la columna de otro dataframe basado en una columna comun como index
    df2 a df1 el nombre de las columnas a actuliazar debe ser el mismo, asi como el de los indexs"""
    df1['original_index'] = df1.index  # Guarda el índice original
    df2_selected = df2[[index_name] + columns]
    df1 = df1.merge(df2_selected, on=index_name, how='left', suffixes=('', '_y'))

    for column in columns:
        df1[column].update(df1.pop(column + '_y'))

    df1.set_index('original_index', inplace=True)
    
    df1 = df1.rename_axis(None)# Restablece el índice original sin nombre 
    df1 = df1.drop_duplicates()
    return df1
def update_values_optimized_V2(df1, df2, index_name, columns):
    """ Implementa la funcionalidad de si el valor dentro de sa columna ya tiene valor, no es afectad"""
    df1['original_index'] = df1.index  # Guarda el índice original
    df2_selected = df2[[index_name] + columns]
    df1 = df1.merge(df2_selected, on=index_name, how='left', suffixes=('', '_y'))

    for column in columns:
        ## el valor es el mismo valor si en la data par actualizar es null, sino se toma el valor de la data a actualizar
        df1[column] = df1.apply(lambda row: row[column] if pd.isnull(row[column + '_y']) else row[column + '_y'], axis=1)
        df1.drop(column + '_y', axis=1, inplace=True)

    df1.set_index('original_index', inplace=True)
    df1 = df1.rename_axis(None)  # Restablece el índice original sin nombre 
    df1 = df1.drop_duplicates()
    return df1

def Excel_format(df, Today_path, useless_columns):
    
    # Exporto a excel
    df.to_excel(Today_path, index=False)
    
    # Load Excel
    wb = load_workbook(Today_path)
    ws = wb.active
    
    # Get DataFrame dimensions
    num_rows = len(df)
    num_cols = len(df.columns)
    
    # Create a table in Excel with the DataFrame data
    tab = Table(displayName="Tabla1", ref="$A$1:${}${}".format(get_column_letter(num_cols), num_rows + 1))
    
    # Hide columns by name
    for col_name in useless_columns:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name)
            ws.column_dimensions[get_column_letter(col_idx + 1)].hidden = True
    
    # Style of table
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(Today_path)
    
#Prepa_EI.loc[:, "OC Posición"] = Prepa_EI.loc[:, "DOC_COMPRAS"].str.cat(Prepa_EI.loc[:, "POSIC"], sep= ":")

## Dont used ###
def pre_proces(df,columns_2str,column_filter):
        ## Preprosecing of PREP_NEW
        df = Ac.convert_columns_to_str(df.copy(),columns_2str ) #Convert to str a key column
        df_EI = df.loc[df[column_filter] == 'Eduardo Iberico']#Filter
        df_EI = df_EI.copy()  # Crea una copia del DataFrame original para evitar modificar los datos originales
        return df_EI



def convert_columns_to_str(df, columns):
    for column in columns:
        df[column] = df[column].astype(str)
    return df

"""
    df1 es mas grande en filas que df2 siempre
    de otra forma los index se """
def merge_dataframes(df1, df2, merge_on, index):
    merged_df = pd.merge(df1, df2, on=merge_on, how='left')
    merged_df = merged_df.drop_duplicates()
    #print("filas originales: ",len(merged_df))
    #print("filas index: ",len(index))
    merged_df.set_index(index, inplace=True)
    return merged_df
    
def date_format(df, columnas):
    for col in columnas:
        try:
            df[col] = pd.to_datetime(df[col])
            df[col] = df[col].dt.strftime('%d/%m/%Y')
        except ValueError as e:
            print(f"Error al convertir la columna {col}: {e}")
    return df
def crear_area(df):
        map_Analistas = {
        'ANGGIE': 'Construccion',
        'DEMETRIO' : 'Construccion',
        'EDWIN' : 'Construccion',
        'JHORDAN' :'Construccion',
        'JORGE' : 'Construccion',
        'Marlon' : 'Construccion',
        'JENNY' : 'Implementacion',
        'DANNER' : 'Implementacion',
        'MARCO' : 'Implementacion',
        'LAURA' : 'Implementacion'}
        df['Area'] = df['RESPONSABLE EA'].map(map_Analistas)
        return df 

def melt_columns(df,columna):
    df_2slipt= df.copy() 
    df_split = df_2slipt[columna].str.split('/', expand=True)
    # Luego, usamos stack para "despivotar" las columnas generadas "POS"
    #Borro el el subindex generado y lo reseteo. Para que cada POS Coincida con su index original 
    df_split_stacked = df_split.stack().reset_index(level=1, drop=True).rename('Posicion')
    # Finalmente, unimos el DataFrame original con la serie "despivotada", Se une por el index, mediante una cosulta que es la
    ## duplica las filas, Realmente solo se splitean las ocs y se conserva el index.
    df_final= df.join(df_split_stacked)
    df_final.drop(columns = columna, inplace= True) 
    return df_final
        

def split_ocs(PAP):
    """ Funcion que splitea las OCs multiples, las normaliza y las vuelve a unir en un formato estandar para todas las OCs
        Esto me ayuda al aplicar el algoritmo de actualizacion de data"""
    
    PAP_f = PAP.dropna(subset=['OC Posición'])
    PAP_f_Multi = PAP_f[PAP_f['OC Posición'].str.contains('/')]
    PAP_f_Mono = PAP_f[~PAP_f['OC Posición'].str.contains('/')]
    mask = PAP_f_Multi['OC Posición'].str.contains(r'\d+:\d+/\d+:\d+')
    ## Dentro de las Ocs Multi hay 2 CLases A & B 
    PAP_f_Multi_A = PAP_f_Multi[mask]  # Contiene las filas con el formato "*:*/*:*"
    PAP_f_Multi_B = PAP_f_Multi[~mask]
    #Proceso las clase A
    PAP_f_Multi_A_final = melt_columns(PAP_f_Multi_A,'OC Posición')
    PAP_f_Multi_A_final.rename(columns={'Posicion': 'OC Posición'},inplace= True)
    #Pre-Proceso la clase B
    PAP_f_Multi_B = PAP_f_Multi_B.copy()
    # Separar la columna 'OC Posición' en varias columnas
    split_data = PAP_f_Multi_B['OC Posición'].str.split(':', expand=True)
    # Renombrar las columnas con un prefijo y el número de columna
    split_data.columns = [f"new_column_{i}" for i in range(split_data.shape[1])]
    # Unir las nuevas columnas al DataFrame original
    PAP_f_Multi_B = PAP_f_Multi_B.join(split_data)
    # Eliminar las columnas que ya no necesitas
    PAP_f_Multi_B = PAP_f_Multi_B.iloc[:, :12]
    PAP_f_Multi_B = PAP_f_Multi_B.rename(columns={'new_column_0':'OC',
                                                  'new_column_1':'POS'})
    PAP_f_Multi_B = PAP_f_Multi_B.drop('OC Posición', axis=1)
    #Proceso la clase B
    PAP_f_Multi_B_final = melt_columns(PAP_f_Multi_B,'POS')
    PAP_f_Multi_B_final = convert_columns_to_str(PAP_f_Multi_B_final.copy(), ['OC', 'Posicion']) #Convert to str a key column
    PAP_f_Multi_B_final["OC Posición"] = PAP_f_Multi_B_final["OC"].str.cat(PAP_f_Multi_B_final["Posicion"], sep= ":") # format
    PAP_f_Multi_B_final.drop(columns = ['OC','Posicion'],inplace=True)
    df_concat = pd.concat([PAP_f_Multi_B_final, PAP_f_Multi_A_final]) 
    df_concat_final = pd.concat([df_concat, PAP_f_Mono]) 
    
    return df_concat_final 

def equal_columns(df1,df2):
    """ Para que df2 tenga las columnas que le faltan de df1, usado para el pre-merge"""
    a = set(df1.columns.tolist()) - set(df2.columns.tolist()) # Columnas que le faltan al grande del chico
    # Crear las columnas necesarias con dtype 'object'
    for col in list(a):
        df2[col] = pd.Series(dtype='object')

def addResp(EA_act: pd.DataFrame,Pext_N: pd.DataFrame):
    EA_act.loc[:,'original_index'] = EA_act.index
    EA_act['SITE'] = EA_act['SITE'].str.replace('_', ' ')
    Pext_N['SITE'] = Pext_N['SITE'].str.replace('_', ' ')
    EA_act['SITE'] = EA_act['SITE'].str.replace('Ã`', 'Ñ')
    Pext_N['SITE'] = Pext_N['SITE'].str.replace('Ã`', 'Ñ')
    EA_act_df = pd.merge(EA_act,Pext_N,on=['SITE','PROVEEDOR'], how='left') 
    EA_act_df= EA_act_df.drop_duplicates()
    EA_act_df.loc[EA_act_df['RESPONSABLE EA'].isna(), 'RESPONSABLE EA'] = EA_act_df.loc[EA_act_df['RESPONSABLE EA'].isna(), 'RESPONSABLE DE IMPLEMENTACION']
    df_duplicate = EA_act_df[EA_act_df.duplicated(subset='original_index',keep=False)].sort_values(by='original_index')
    df_noduplicate = EA_act_df[~EA_act_df.duplicated(subset='original_index',keep=False)].sort_values(by='original_index')# Separo los duplicados y los que no
    
    df_duplicate['DESCRIPCIONPROY'] = df_duplicate['DESCRIPCIONPROY'].str.strip() # quito espacios a esas 2 columnas para el df de duplicados
    df_duplicate['PROYECTO'] = df_duplicate['PROYECTO'].str.strip()
    df = df_duplicate[df_duplicate['PROYECTO'] == df_duplicate['DESCRIPCIONPROY']]
    EA_act_df = pd.concat([df_noduplicate,df])
    EA_act_df.set_index('original_index', inplace=True)
    EA_act_df = EA_act_df.rename_axis(None)# Restablece el índice original sin nombre 
    EA_act_df = EA_act_df.drop(columns=['RESPONSABLE DE IMPLEMENTACION','PROYECTO'])
    EA_act_df['RESPONSABLE EA'] = EA_act_df['RESPONSABLE EA'].replace({'Laura Rafael' : "LAURA" ,
                                                                       'Danner Yarleque' : 'DANNER',
                                                                       'Jenny Pizan' : 'JENNY'})
                                                                      
    return EA_act_df

def addSite(EA_act_df : pd.DataFrame, SAP_4_use : pd.DataFrame):
    """Funcione que toma las tablas SAP y extrae los sites y los codigos de AHI
        Devuelve un DF con la Informacion Añadida ahi"""
    df_merged = merge_dataframes(EA_act_df,SAP_4_use,'CONCATENADO',EA_act_df.index)
    df_merged = df_merged.drop_duplicates()
    df_merged['ID_SITE_SAP'] = df_merged['PEP Desc'].str.slice(start=0 ,stop=6)# Creo la columna de codigo de Site
    df_merged['SITE_SAP'] = df_merged['PEP Desc'].str.slice(start=7)# Creo la columna con el nombre del SITE 

    # Verifica si las columnas 'SITE' e 'ID_SITIO' existen, si no, las crea
    if 'SITE' not in df_merged.columns:
        df_merged['SITE'] = np.nan
    if 'ID_SITIO' not in df_merged.columns:
        df_merged['ID_SITIO'] = np.nan

    # Reemplaza los valores NaN en 'SITE' e 'ID_SITIO' con los valores de 'SITE_SAP' e 'ID_SITE_SAP'
    df_merged['SITE'] = df_merged['SITE'].combine_first(df_merged['SITE_SAP'])
    df_merged['ID_SITIO'] = df_merged['ID_SITIO'].combine_first(df_merged['ID_SITE_SAP'])

    df_merged.SITE = df_merged.SITE.str.strip() # Quito espacios

    EA_act_df = df_merged.drop(columns=['ID_SITE_SAP','SITE_SAP','Fecha OC']) # Elimino las columnas que use para el Merge 
    return EA_act_df

def process_file(filename, sheet_name='Sheet1'):
    df = pd.read_excel(filename, sheet_name=sheet_name, dtype={'COMENTARIO': str})
    df = df[df.RESPONSABLE2 == 'Eduardo Iberico'] #Filtro 
    return df

def calculate_difference(df1, df2):
    pivot_table1 = pd.pivot_table(df1, values='PEND_FACT_SOLES', index=['NOMPROVEEDOR', 'ESTADO'], aggfunc=pd.Series.sum)
    pivot_table2 = pd.pivot_table(df2, values='PEND_FACT_SOLES', index=['NOMPROVEEDOR', 'ESTADO'], aggfunc=pd.Series.sum)

    all_rows = set(pivot_table1.index).union(set(pivot_table2.index))

    pivot_table1 = pivot_table1.reindex(all_rows, fill_value=0)
    pivot_table2 = pivot_table2.reindex(all_rows, fill_value=0)

    diff_pivot_table = pivot_table1 - pivot_table2
    diff_pivot_table = diff_pivot_table.fillna(0)
    diff_pivot_table_reset = diff_pivot_table.reset_index()
    diff_pivot_table_reset.rename(columns={'PEND_FACT_SOLES': 'Dif Soles'}, inplace=True)
    # Fusiona diff_pivot_table_filled_reset con PrePa_O_EI
    PrePa_O_EI = pd.merge(df1, diff_pivot_table_reset, on=['NOMPROVEEDOR', 'ESTADO'], how='left')

    return PrePa_O_EI
    
def CalcDiff(Week_Int):
    path = 'D:/Scripts1/Code/ActPEA/Resultados_prepa/*'
    tipo_de_archivo = '*.xlsx'
    archivos = glob.glob(path + tipo_de_archivo)
    
    archivos_ordenados = sorted(archivos, key=os.path.getctime, reverse=True)
     
    nombres_de_archivos = [os.path.basename(archivo) for archivo in archivos_ordenados[:1] + archivos_ordenados[Week_Int-1:Week_Int]] 
    
    print(nombres_de_archivos)
    
    dataframes = [process_file(os.path.join(path[:-1], nombre)) for nombre in nombres_de_archivos]
    
    diff_pivot_table_reset = calculate_difference(dataframes[0], dataframes[1])
    return diff_pivot_table_reset
def convert_to_date(val):
    try:
        return pd.Timestamp('1899-12-30') + pd.Timedelta(int(float(val)), 'D')
    except ValueError:
        return val
def get_recent_df(Carpeta_path: str,sheet_name: str):
    """ Devuelve el df de la hoja especifica, del archivo mas reciente de la carpeta especificada"""
    Path_n= Carpeta_path + '/*'
    tipo_de_archivo = '*.xlsx'
    # Busca el archivo más reciente
    archivos = glob.glob(Path_n + tipo_de_archivo)
    archivo_mas_reciente = max(archivos, key=os.path.getctime)
    nombre_del_archivo_N = os.path.basename(archivo_mas_reciente)
    print(archivo_mas_reciente)
    df = pd.read_excel(archivo_mas_reciente , sheet_name=sheet_name) 
    return df
